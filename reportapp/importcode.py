import os
import sys
import django
from django.db import transaction
from openpyxl import load_workbook

# Project root qo‘shamiz
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(BASE_DIR)

# Django environment setup
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "new_york_taxi.settings")
django.setup()

from reportapp.models import TaxiTrip


def import_excel_full(file_path, batch_size=5000):
    """
    Imports a large Excel file with full columns and computes 'category' field.
    Memory-efficient using openpyxl read_only.
    """
    if not os.path.exists(file_path):
        print(f"❌ Xatolik: '{file_path}' fayli topilmadi.")
        return

    print("Excel faylini o'qish...")

    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active

        batch_records = []
        created = 0

        # Header row
        headers = [cell.value for cell in sheet[1]]

        # Index mapping
        col_idx = {col: idx for idx, col in enumerate(headers)}

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            distance = row[col_idx['trip_distance']]

            # Category
            if distance < 1:
                trip_type = 1
            elif distance < 5:
                trip_type = 2
            elif distance < 10:
                trip_type = 3
            else:
                trip_type = 4

            # Model instance yaratish
            batch_records.append(TaxiTrip(
                vendor_id=row[col_idx.get('vendor_id')],
                pickup_datetime=row[col_idx.get('pickup_datetime')],
                dropoff_datetime=row[col_idx.get('dropoff_datetime')],
                passenger_count=row[col_idx.get('passenger_count')],
                trip_distance=distance,
                pickup_longitude=row[col_idx.get('pickup_longitude')],
                pickup_latitude=row[col_idx.get('pickup_latitude')],
                rate_code=row[col_idx.get('rate_code')],
                store_and_fwd_flag=row[col_idx.get('store_and_fwd_flag')],
                dropoff_longitude=row[col_idx.get('dropoff_longitude')],
                dropoff_latitude=row[col_idx.get('dropoff_latitude')],
                payment_type=row[col_idx.get('payment_type')],
                fare_amount=row[col_idx.get('fare_amount')],
                surcharge=row[col_idx.get('surcharge')],
                mta_tax=row[col_idx.get('mta_tax')],
                tip_amount=row[col_idx.get('tip_amount')],
                tolls_amount=row[col_idx.get('tolls_amount')],
                total_amount=row[col_idx.get('total_amount')],
                category=trip_type
            ))

            # Bulk create
            if len(batch_records) >= batch_size:
                with transaction.atomic():
                    TaxiTrip.objects.bulk_create(batch_records, batch_size)
                created += len(batch_records)
                print(f"{created} ta row import qilindi...")
                batch_records = []

        # Oxirgi qolganlar
        if batch_records:
            with transaction.atomic():
                TaxiTrip.objects.bulk_create(batch_records, batch_size)
            created += len(batch_records)

        print(f"✅ Import tugadi. Jami {created} ta row import qilindi.")

    except Exception as e:
        print(f"Importda error: {e}")

if __name__ == '__main__':
    excel_file_path = "/home/asus/Downloads/Telegram Desktop/new_york_taxi_1_mln.xlsx"
    import_excel_full(excel_file_path)