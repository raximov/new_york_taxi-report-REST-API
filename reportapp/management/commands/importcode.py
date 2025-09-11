import os
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from reportapp.models import TaxiTrip


class Command(BaseCommand):
    help = "Import large Excel file into TaxiTrip model (with category calculation)"

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            type=str,
            help="Path to the Excel file (.xlsx)",
        )
        parser.add_argument(
            "--batch-size",
            type=int,
            default=100000,
            help="Number of rows per bulk insert (default: 100 000)",
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]
        batch_size = options["batch_size"]

        if not os.path.exists(file_path):
            self.stderr.write(f" Xatolik: '{file_path}' fayli topilmadi.")
            return

        self.stdout.write("Excel faylini o‘qish boshlandi...")

        try:
            workbook = load_workbook(filename=file_path, read_only=True)
            sheet = workbook.active

            batch_records = []
            created = 0

            # Header row
            headers = [cell.value for cell in sheet[1]]
            col_idx = {col: idx for idx, col in enumerate(headers)}

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                distance = row[col_idx['trip_distance']]

                # Category hisoblash
                if distance < 1:
                    trip_type = 1
                elif distance < 5:
                    trip_type = 2
                elif distance < 10:
                    trip_type = 3
                else:
                    trip_type = 4

                batch_records.append(TaxiTrip(
                    vendor_id=row[col_idx.get('vendor_id')],
                    pickup_datetime=row[col_idx.get('pickup_datetime')],
                    dropoff_datetime=row[col_idx.get('dropoff_datetime')],
                    passenger_count=row[col_idx.get('passenger_count')],
                    trip_distance=distance,
                    pickup_longitude=row[col_idx.get('pickup_longitude')],
                    pickup_latitude=row[col_idx.get('pickup_latitude')],
                    rate_code=row[col_idx.get('rate_code')],
                    store_and_fwd_flag=row[col_idx.get('store_and_fwd_flag')],
                    dropoff_longitude=row[col_idx.get('dropoff_longitude')],
                    dropoff_latitude=row[col_idx.get('dropoff_latitude')],
                    payment_type=row[col_idx.get('payment_type')],
                    fare_amount=row[col_idx.get('fare_amount')],
                    surcharge=row[col_idx.get('surcharge')],
                    mta_tax=row[col_idx.get('mta_tax')],
                    tip_amount=row[col_idx.get('tip_amount')],
                    tolls_amount=row[col_idx.get('tolls_amount')],
                    total_amount=row[col_idx.get('total_amount')],
                    category=trip_type
                ))

                # Bulk create
                if len(batch_records) >= batch_size:
                    with transaction.atomic():
                        TaxiTrip.objects.bulk_create(batch_records, batch_size)
                    created += len(batch_records)
                    self.stdout.write(f" {created} ta row import qilindi...")
                    batch_records = []

            # Oxirgi qolganlar
            if batch_records:
                with transaction.atomic():
                    TaxiTrip.objects.bulk_create(batch_records, batch_size)
                created += len(batch_records)

            self.stdout.write(self.style.SUCCESS(
                f"🎉 Import tugadi. Jami {created} ta row import qilindi."
            ))

        except Exception as e:
            self.stderr.write(f" Importda error: {e}")
